"""Shared XML renderer for spreadsheet text extraction.

Both the values (data_only=True) and formulas (data_only=False) extractors
use this same format so the LLM judge can cross-reference cell-by-cell.
"""

from pathlib import Path

import openpyxl


def _col_letter(col_idx: int) -> str:
    result = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _xml_escape(val: str) -> str:
    return (
        val.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def spreadsheet_to_xml(file_path: Path, *, data_only: bool) -> str:
    read_only = data_only
    wb = openpyxl.load_workbook(
        str(file_path), data_only=data_only, read_only=read_only
    )
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_xml: list[str] = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
                cells_xml: list[str] = []
                has_value = False
                for cell in row:
                    col = (
                        _col_letter(cell.column)
                        if hasattr(cell, "column") and cell.column is not None
                        else _col_letter(len(cells_xml) + 1)
                    )
                    ref = f"{col}{row_idx}"
                    val = cell.value
                    if val is None:
                        cells_xml.append(f'    <cell ref="{ref}" />')
                    else:
                        has_value = True
                        escaped = _xml_escape(str(val))
                        if isinstance(val, str) and not str(val).startswith("="):
                            cells_xml.append(
                                f'    <cell ref="{ref}" type="string">{escaped}</cell>'
                            )
                        else:
                            cells_xml.append(f'    <cell ref="{ref}">{escaped}</cell>')
                if has_value:
                    rows_xml.append(
                        f'  <row number="{row_idx}">\n'
                        + "\n".join(cells_xml)
                        + "\n  </row>"
                    )
            if rows_xml:
                escaped_name = _xml_escape(sheet_name)
                parts.append(
                    f'<template sheet="{escaped_name}">\n'
                    + "\n".join(rows_xml)
                    + "\n</template>"
                )
    finally:
        wb.close()
    return "\n\n".join(parts)
