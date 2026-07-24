"""File utilities for spreadsheet verifier."""

import csv
import fnmatch
import io
import os
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension as _ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet

from .cell_utils import SpreadsheetData
from .config import EXCEL_EXTENSIONS, SUPPORTED_EXTENSIONS

# Monkey-patch openpyxl ColumnDimension to accept 'level' kwarg.
# openpyxl 3.1.x has a bug where its XML parser passes 'level' (from <col level="1"/>)
# to ColumnDimension.__init__(), but the class only accepts 'outlineLevel'.
# This causes TypeError when loading Excel files with column outline/grouping levels.
_original_cd_init = _ColumnDimension.__init__


def _patched_cd_init(self, *args, level=None, **kwargs):  # type: ignore[no-untyped-def]
    if level is not None and "outlineLevel" not in kwargs:
        kwargs["outlineLevel"] = level
    _original_cd_init(self, *args, **kwargs)


_ColumnDimension.__init__ = _patched_cd_init  # type: ignore[method-assign]

# Pattern for simple cell reference formulas: =A1, =Sheet1!B2, ='Sheet Name'!C3
_SIMPLE_REF_PATTERN = re.compile(
    r"^=(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?"
    r"([A-Za-z]{1,3})(\d+)$"
)


def find_matching_files(zip_file: zipfile.ZipFile, file_pattern: str) -> list[str]:
    """Find files in zip matching the pattern.

    Pattern can be:
    - Exact filename: "data.csv"
    - Extension pattern: "*.csv"
    - Path pattern: "output/*.xlsx"
    """
    matching = []
    for name in zip_file.namelist():
        if name.endswith("/"):
            continue

        normalized = name.replace("\\", "/").strip("/")
        basename = Path(normalized).name

        if fnmatch.fnmatch(basename, file_pattern) or fnmatch.fnmatch(
            normalized, file_pattern
        ):
            ext = Path(name).suffix.lower()
            if ext in SUPPORTED_EXTENSIONS:
                matching.append(name)

    return matching


def load_spreadsheet_data(
    file_bytes: io.BytesIO, file_ext: str, sheet_name: str | None = None
) -> SpreadsheetData:
    """Load a spreadsheet file into SpreadsheetData.

    Data is loaded without treating the first row as headers,
    so cell references like A1, B2 map directly to row/column indices.
    """
    file_bytes.seek(0)

    if file_ext == ".csv":
        return _load_csv(file_bytes)

    file_bytes.seek(0)
    return _load_excel(file_bytes, sheet_name)


def _load_csv(file_bytes: io.BytesIO) -> SpreadsheetData:
    """Load CSV file into SpreadsheetData."""
    for encoding in ["utf-8", "latin-1", "cp1252"]:
        try:
            file_bytes.seek(0)
            text = file_bytes.read().decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        file_bytes.seek(0)
        text = file_bytes.read().decode("utf-8", errors="replace")

    rows: list[list[Any]] = []
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        converted_row: list[Any] = []
        for cell in row:
            converted_row.append(_convert_cell_value(cell))
        rows.append(converted_row)

    return SpreadsheetData(rows=rows)


def _load_excel(
    file_bytes: io.BytesIO, sheet_name: str | None = None
) -> SpreadsheetData:
    """Load Excel file into SpreadsheetData using openpyxl."""
    file_bytes.seek(0)
    wb = load_workbook(file_bytes, data_only=True, read_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return SpreadsheetData(rows=[])
            ws = wb[sheet_name]
        else:
            if not wb.worksheets:
                return SpreadsheetData(rows=[])
            ws = wb.worksheets[0]

        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
    finally:
        wb.close()

    # Quick check: if no None values, no formulas to resolve
    if not any(val is None for row in rows for val in row):
        return SpreadsheetData(rows=rows)

    # Attempt to resolve formula cells that returned None cached values
    _try_resolve_formulas(file_bytes, rows, sheet_name)

    return SpreadsheetData(rows=rows)


def _try_resolve_formulas(
    file_bytes: io.BytesIO,
    rows: list[list[Any]],
    sheet_name: str | None,
) -> None:
    """Detect and resolve formula cells that have no cached values.

    When Excel files are saved by tools that don't compute formulas (openpyxl,
    LibreOffice, etc.), formula cells have no cached result and data_only=True
    returns None.

    Two-pass resolution:
    1. Simple reference resolution — handles =Sheet!A1, =A1 (fast, no extra deps)
    2. Full formula evaluation via `formulas` library — handles =B2*B4, =SUM(...), etc.
    """
    file_bytes.seek(0)
    wb = load_workbook(file_bytes, data_only=False)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]
        else:
            if not wb.worksheets:
                return
            ws = wb.worksheets[0]

        current_sheet_name = ws.title

        # Collect formula cells that have no cached value
        formula_cells: list[tuple[int, int]] = []
        for row_idx, row in enumerate(rows):
            for col_idx, val in enumerate(row):
                if val is not None:
                    continue
                cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                cell_value = cell.value
                if cell_value is not None and str(cell_value).startswith("="):
                    formula_cells.append((row_idx, col_idx))

        if not formula_cells:
            return

        # Pass 1: Simple reference resolution (fast, handles =Sheet!A1)
        simple_resolved = 0
        for row_idx, col_idx in formula_cells:
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
            resolved = _resolve_formula_reference(
                str(cell.value), wb, current_sheet_name
            )
            if resolved is not None:
                rows[row_idx][col_idx] = resolved
                simple_resolved += 1

        if simple_resolved > 0:
            logger.info(
                f"[SPREADSHEET_VERIFIER] Resolved {simple_resolved} formula cell(s) "
                f"via simple reference lookup"
            )

        # Check if any formulas remain unresolved
        remaining = [(r, c) for r, c in formula_cells if rows[r][c] is None]
        if not remaining:
            return

        # Pass 2: Full formula evaluation with formulas library
        _try_evaluate_with_formulas_lib(file_bytes, rows, remaining, sheet_name)
    finally:
        wb.close()


def _try_evaluate_with_formulas_lib(
    file_bytes: io.BytesIO,
    rows: list[list[Any]],
    formula_cells: list[tuple[int, int]],
    sheet_name: str | None,
) -> None:
    """Evaluate complex formulas using the `formulas` library.

    Writes the workbook to a temp file, computes all formulas via
    formulas.ExcelModel, then reads back the computed values.

    Falls back silently if evaluation fails for a specific formula.
    """
    import formulas as formulas_lib  # pyright: ignore[reportMissingTypeStubs]

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            file_bytes.seek(0)
            tmp.write(file_bytes.read())
            temp_path = tmp.name

        xl_model = formulas_lib.ExcelModel().loads(temp_path).finish()
        xl_model.calculate()

        output_dir = tempfile.mkdtemp()
        try:
            xl_model.write(dirpath=output_dir)

            # The formulas library uppercases the book name internally,
            # so the output filename may not match the original temp_path
            # basename. Find the actual .xlsx file in the output directory.
            xlsx_files = [
                f
                for f in os.listdir(output_dir)
                if f.endswith(".xlsx") or f.endswith(".XLSX")
            ]
            if not xlsx_files:
                return
            output_path = os.path.join(output_dir, xlsx_files[0])

            wb_computed = load_workbook(output_path, data_only=True)
            try:
                if sheet_name:
                    # The formulas library also uppercases sheet names,
                    # so do a case-insensitive lookup.
                    matched_sheet = None
                    for s in wb_computed.sheetnames:
                        if s.lower() == sheet_name.lower():
                            matched_sheet = s
                            break
                    if matched_sheet is None:
                        return
                    ws = wb_computed[matched_sheet]
                else:
                    if not wb_computed.worksheets:
                        return
                    ws = wb_computed.worksheets[0]

                resolved_count = 0
                for row_idx, col_idx in formula_cells:
                    if rows[row_idx][col_idx] is not None:
                        continue
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                    if cell.value is not None:
                        rows[row_idx][col_idx] = cell.value
                        resolved_count += 1

                if resolved_count > 0:
                    logger.info(
                        f"[SPREADSHEET_VERIFIER] Evaluated {resolved_count} "
                        f"formula cell(s) via formulas library"
                    )
            finally:
                wb_computed.close()
        finally:
            # Clean up output directory
            for f in os.listdir(output_dir):
                os.unlink(os.path.join(output_dir, f))
            os.rmdir(output_dir)
    except Exception as e:
        logger.debug(f"[SPREADSHEET_VERIFIER] formulas library evaluation failed: {e}")
    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


def _resolve_formula_reference(
    formula: str, wb: Workbook, current_sheet_name: str
) -> Any:
    """Resolve a simple cell reference formula like =Sheet1!A1 or =A1.

    Handles:
    - =A1 (same-sheet reference)
    - =Sheet1!B2 (cross-sheet reference)
    - ='Sheet Name'!C3 (quoted sheet name)

    Complex formulas (SUM, IF, arithmetic, etc.) are not resolved.

    Returns the resolved value or None if not resolvable.
    """
    match = _SIMPLE_REF_PATTERN.match(formula.strip())
    if not match:
        return None

    ref_sheet = match.group(1) or match.group(2)
    col = match.group(3)
    row = int(match.group(4))
    target_sheet = ref_sheet or current_sheet_name

    try:
        if target_sheet not in wb.sheetnames:
            return None
        ws = wb[target_sheet]
        value = ws[f"{col}{row}"].value
        # If the referenced cell is also a formula, we can't resolve further
        if value is not None and str(value).startswith("="):
            return None
        return value
    except (KeyError, ValueError, IndexError):
        return None


def _convert_cell_value(value: str) -> Any:
    """Convert string cell value to appropriate type."""
    if not value:
        return None

    try:
        return int(value)
    except ValueError:
        pass

    try:
        return float(value)
    except ValueError:
        pass

    return value


def load_workbook_from_bytes(file_bytes: bytes) -> Workbook:
    """Load an openpyxl Workbook from bytes."""
    return load_workbook(io.BytesIO(file_bytes), data_only=False)


def get_worksheet(
    workbook: Workbook, sheet_name: str | None = None
) -> tuple[Worksheet | None, str | None]:
    """Get worksheet by name or first sheet.

    Returns:
        Tuple of (worksheet, error_message). If error, worksheet is None.
    """
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            return (
                None,
                f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}",
            )
        return workbook[sheet_name], None

    # Use first sheet by index for consistent, predictable behavior
    # (workbook.active returns whichever sheet was last viewed, not necessarily first)
    if not workbook.worksheets:
        return None, "No worksheets found in workbook"
    return workbook.worksheets[0], None


def is_excel_file(file_ext: str) -> bool:
    """Check if file extension is Excel."""
    return file_ext.lower() in EXCEL_EXTENSIONS
