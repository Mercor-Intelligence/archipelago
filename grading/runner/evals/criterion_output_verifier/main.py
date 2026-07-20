"""Criterion Output Verifier

combines apex v1 rubric structure with output_llm grading

Field mappings from the criterion schema to output_llm:
- description -> criteria
- rationale -> included in grading context
- sources -> artifacts_to_reference
- expected_file_type -> expected_file_type
- convert_to_images -> converts output artifacts to images for VLM formatting checks

Additional criterion metadata (passthrough only, not used in grading):
- type: "Critical", "Nice-to-have", "Style / Formatting", "Subjective"
- tags: ["Extraction", "Reasoning", "Style"]
"""

import io
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
from litellm import Choices
from loguru import logger

from runner.evals.models import EvalImplInput
from runner.evals.output_llm.artifact_filters import (
    artifact_matches_filters,
    convert_file_types_to_extensions,
    is_valid_file_type,
    should_filter_all_files,
    should_skip_filter,
)
from runner.evals.output_llm.utils.log_helpers import (
    get_artifact_identity,
    log_artifact_filter,
    log_artifact_selector_result,
    log_diff_extraction,
    log_grader_final_prompt,
    log_grader_result,
    log_grader_start,
    log_grader_truncation,
)
from runner.evals.output_llm.utils.prompts import (
    GRADING_SYSTEM_PROMPT,
    GRADING_SYSTEM_PROMPT_NO_REFERENCE,
    GradingResponseSchema,
)
from runner.evals.output_llm.utils.services.artifact_evaluate import (
    select_artifacts_to_evaluate,
)
from runner.evals.output_llm.utils.services.prompt_builder import build_grading_prompt
from runner.evals.output_llm_multi_representation.artifact_representations import (
    ArtifactSelectionWithTransformations,
    fetch_artifacts_with_transformations,
)
from runner.helpers.models import HelperIds
from runner.helpers.snapshot_diff import extract_artifact_changes_from_diff
from runner.models import VerifierResult
from runner.utils.file_extraction.utils.chart_extraction import (
    evaluate_excel_formulas_with_libreoffice,
)
from runner.utils.file_transformations.docx_to_images.main import docx_to_images
from runner.utils.file_transformations.pptx_to_images.main import pptx_to_images
from runner.utils.file_transformations.spreadsheet_to_images.main import (
    spreadsheet_to_images,
)
from runner.utils.llm import build_messages, call_llm
from runner.utils.token_utils import count_tokens, get_model_context_limit

LLM_JUDGE_TIMEOUT = 3600

# File extension -> image conversion function
_IMAGE_CONVERTERS = {
    ".docx": docx_to_images,
    ".pptx": pptx_to_images,
    ".xlsx": spreadsheet_to_images,
    ".xlsm": spreadsheet_to_images,
}

_XLSX_EXTENSIONS = {".xlsx", ".xlsm"}


class _SheetFormulaInfo:
    """Formula map, computed values, and dimensions for a single sheet."""

    __slots__ = ("formulas", "computed_values", "max_row", "max_col")

    def __init__(
        self,
        formulas: dict[tuple[int, int], str],
        computed_values: dict[tuple[int, int], str],
        max_row: int,
        max_col: int,
    ):
        self.formulas = formulas
        self.computed_values = computed_values
        self.max_row = max_row
        self.max_col = max_col


def _build_formula_map(
    file_bytes: bytes,
) -> dict[str, _SheetFormulaInfo]:
    """Extract formulas from an xlsx (data_only=False).

    Returns {sheet_name: _SheetFormulaInfo} for sheets that contain formulas.
    The ``computed_values`` dict is initially empty; the caller populates it
    by running LibreOffice evaluation.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    result: dict[str, _SheetFormulaInfo] = {}
    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            fmap: dict[tuple[int, int], str] = {}
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.row is None or cell.column is None:
                        continue
                    if cell.data_type == "f" or (
                        cell.value is not None and str(cell.value).startswith("=")
                    ):
                        formula = str(cell.value)
                        if not formula.startswith("="):
                            formula = f"={formula}"
                        fmap[(cell.row, cell.column)] = formula
            if fmap:
                result[sheet_name] = _SheetFormulaInfo(
                    formulas=fmap,
                    computed_values={},
                    max_row=sheet.max_row or 0,
                    max_col=sheet.max_column or 0,
                )
    finally:
        wb.close()
    return result


def _rebuild_sheet_content_with_formulas(
    sheet_name: str,
    info: "_SheetFormulaInfo",
    evaluated_wb: "openpyxl.Workbook | None",
) -> str:
    """Rebuild sheet text from the evaluated workbook, annotating formula cells.

    Instead of trying to patch the existing tab-separated text (which loses
    row/column coordinate mapping due to empty-row/cell filtering), we rebuild
    the content from the evaluated workbook directly so coordinates are exact.

    Output format matches the extractor: ``=== Sheet: Name ===`` header,
    followed by tab-separated rows with formula cells annotated as
    ``value (=FORMULA)`` or ``(=FORMULA)`` when no value is available.
    """
    lines = [f"=== Sheet: {sheet_name} ==="]

    if evaluated_wb and sheet_name in evaluated_wb.sheetnames:
        sheet = evaluated_wb[sheet_name]
    else:
        sheet = None

    for row_idx in range(1, info.max_row + 1):
        row_cells: list[str] = []
        has_value = False
        for col_idx in range(1, info.max_col + 1):
            formula = info.formulas.get((row_idx, col_idx))
            computed = info.computed_values.get((row_idx, col_idx))

            # Get the display value from the evaluated workbook
            val: str | None = None
            if sheet is not None:
                cell_val = sheet.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    val = str(cell_val)

            if formula:
                if val:
                    row_cells.append(f"{val} ({formula})")
                elif computed:
                    row_cells.append(f"{computed} ({formula})")
                else:
                    row_cells.append(f"({formula})")
                has_value = True
            elif val:
                row_cells.append(val)
                has_value = True
            else:
                row_cells.append("")

        if has_value:
            # Strip trailing empty cells
            while row_cells and row_cells[-1] == "":
                row_cells.pop()
            if row_cells:
                lines.append("\t".join(row_cells))

    return "\n".join(lines)


async def _annotate_artifacts_with_formulas(
    artifacts: list[Any],
    snapshot_zip: zipfile.ZipFile,
    task_id: str,
) -> None:
    """Annotate xlsx artifact content with formula info in-place.

    For each xlsx sheet artifact, loads the source file to extract formulas,
    evaluates them via LibreOffice for computed values, and rebuilds the
    sheet content with formulas annotated inline: ``16 (=A1+B1)``.

    Only ``new_content`` is rebuilt (from the final snapshot). ``old_content``
    is left unchanged since it represents the pre-change state.
    """
    # Group artifacts by xlsx path (dedup by path)
    xlsx_paths: dict[str, list[Any]] = {}
    for artifact in artifacts:
        ext = Path(artifact.path).suffix.lower()
        if ext in _XLSX_EXTENSIONS and artifact.artifact_type == "sheet":
            xlsx_paths.setdefault(artifact.path, []).append(artifact)

    if not xlsx_paths:
        return

    for xlsx_path, sheet_artifacts in xlsx_paths.items():
        matching = [
            n
            for n in snapshot_zip.namelist()
            if n.endswith(xlsx_path) or xlsx_path in n
        ]
        if not matching:
            continue

        try:
            file_bytes = snapshot_zip.read(matching[0])
            formula_maps = _build_formula_map(file_bytes)
        except Exception as e:
            logger.debug(
                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | Could not extract formulas from {xlsx_path}: {e}"
            )
            continue

        if not formula_maps:
            continue

        # Evaluate formulas via LibreOffice once for the whole file
        evaluated_wb: openpyxl.Workbook | None = None
        evaluated_bytes = await evaluate_excel_formulas_with_libreoffice(file_bytes)
        if evaluated_bytes:
            try:
                evaluated_wb = openpyxl.load_workbook(
                    io.BytesIO(evaluated_bytes), data_only=True
                )
                # Populate computed_values from the evaluated workbook
                for sheet_name, info in formula_maps.items():
                    if sheet_name not in evaluated_wb.sheetnames:
                        continue
                    sheet = evaluated_wb[sheet_name]
                    for row, col in info.formulas:
                        cell_val = sheet.cell(row=row, column=col).value
                        if cell_val is not None:
                            info.computed_values[(row, col)] = str(cell_val)
            except Exception as e:
                logger.debug(
                    f"[JUDGE][CRITERION_OUTPUT] task={task_id} | Failed to load evaluated workbook: {e}"
                )

        annotated_count = 0
        for artifact in sheet_artifacts:
            sheet_name = artifact.title
            info = formula_maps.get(sheet_name)
            if not info:
                continue

            if artifact.new_content:
                artifact.new_content = _rebuild_sheet_content_with_formulas(
                    sheet_name, info, evaluated_wb
                )
                annotated_count += 1

        if evaluated_wb:
            evaluated_wb.close()

        if annotated_count:
            total_formulas = sum(len(i.formulas) for i in formula_maps.values())
            logger.info(
                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | Annotated {annotated_count} sheet(s) in {xlsx_path} with {total_formulas} formula(s)"
            )


def _estimate_artifact_tokens(artifact: Any, model: str) -> int:
    """Estimate tokens for an artifact (same as output_llm)."""
    change_type = artifact.change_type.value

    if change_type == "created":
        content = artifact.new_content or artifact.content_diff or ""
        if content:
            return count_tokens(
                f"<created_content>\n{content}\n</created_content>", model
            )
        return 0

    if change_type == "deleted":
        if artifact.content_diff:
            return count_tokens(
                f"<deleted_content>\n{artifact.content_diff}\n</deleted_content>", model
            )
        return 0

    if change_type == "modified":
        tokens = 0
        if artifact.content_diff:
            tokens += count_tokens(f"<diff>\n{artifact.content_diff}\n</diff>", model)
        if artifact.new_content:
            tokens += count_tokens(
                f"<updated_content>\n{artifact.new_content}\n</updated_content>", model
            )
        return tokens

    if artifact.content_diff:
        return count_tokens(f"<diff>\n{artifact.content_diff}\n</diff>", model)
    return 0


def _extract_task_prompt(input: EvalImplInput) -> str | None:
    """Extract the task prompt from trajectory messages."""
    if not input.trajectory or not input.trajectory.messages:
        return None

    for msg in input.trajectory.messages:
        if msg.get("role") == "user" and msg.get("content"):
            content = msg.get("content")
            return str(content) if content else None

    return None


def _filter_artifacts_programmatically(
    artifacts: list[Any],
    expected_file_type: str,
    task_id: str | None = None,
    criteria: str | None = None,
) -> list[Any]:
    """Filter artifacts by file type (same as output_llm)."""
    if should_filter_all_files(expected_file_type):
        log_artifact_filter(
            task_id or "unknown",
            input_count=len(artifacts),
            output_count=0,
            file_type=expected_file_type,
            filtered_artifacts=[],
            mode="final_answer_only",
            all_artifacts=artifacts,
            criteria=criteria,
        )
        return []

    skip_file_filter = should_skip_filter(expected_file_type)

    if skip_file_filter:
        log_artifact_filter(
            task_id or "unknown",
            input_count=len(artifacts),
            output_count=len(artifacts),
            file_type=expected_file_type,
            filtered_artifacts=artifacts,
            mode="no_filter",
            all_artifacts=artifacts,
            criteria=criteria,
        )
        return artifacts

    allowed_extensions = convert_file_types_to_extensions(expected_file_type)

    filtered = [
        artifact
        for artifact in artifacts
        if artifact_matches_filters(artifact, allowed_extensions)
    ]

    log_artifact_filter(
        task_id or "unknown",
        input_count=len(artifacts),
        output_count=len(filtered),
        file_type=expected_file_type,
        filtered_artifacts=filtered,
        all_artifacts=artifacts,
        criteria=criteria,
    )

    return filtered


def _get_artifact_display_names(artifacts: list[Any]) -> str:
    """Build comma-separated string of artifact display names."""
    if not artifacts:
        return ""

    names = []
    for artifact in artifacts:
        path = artifact.path
        artifact_type = artifact.artifact_type

        if artifact_type in ("slide", "sheet", "page") and artifact.index is not None:
            type_label = artifact_type.capitalize()
            index_display = artifact.index + 1
            if artifact.title:
                names.append(f"{path} ({type_label} {index_display}: {artifact.title})")
            else:
                names.append(f"{path} ({type_label} {index_display})")
        else:
            names.append(path)

    return ", ".join(names)


def _should_auto_fail_missing_file_type(
    expected_file_type: str,
    filtered_artifacts: list[Any],
) -> bool:
    """Check if criterion should auto-fail due to missing file type."""
    if should_skip_filter(expected_file_type):
        return False

    if should_filter_all_files(expected_file_type):
        return False

    return len(filtered_artifacts) == 0


def _build_criterion_prompt(
    description: str,
    rationale: str | None,
) -> str:
    """
    Build the criteria prompt with the criterion's additional context.

    Includes the rationale to help the judge understand why this criterion matters.
    """
    parts = [description]

    if rationale:
        parts.append(f"\n\n[Rationale for this criterion: {rationale}]")

    return "".join(parts)


async def criterion_output_verifier_eval(input: EvalImplInput) -> VerifierResult:
    """
    Grade agent output using the criterion output verifier.

    Uses the same LLM judging logic as output_llm but with the criterion field structure:
    - description: The criterion to evaluate (required)
    - rationale: Why this criterion matters (required)
    - type: Category of the criterion (required, passthrough only)
    - tags: Categorization tags (optional, passthrough only)
    - sources: Reference artifacts for grading (optional)
    - expected_file_type: Grading target (required)
    """
    verifier_values = input.verifier.verifier_values or {}
    task_id = input.verifier.task_id or "unknown"

    description = verifier_values.get("description", "")
    rationale = verifier_values.get("rationale", "")
    verifier_type = verifier_values.get("type", "Critical")
    tags = verifier_values.get("tags") or verifier_values.get("criterion_type", [])
    sources = verifier_values.get("sources", [])
    expected_file_type = verifier_values.get("expected_file_type")
    convert_to_images_raw = verifier_values.get("convert_to_images", False)
    convert_to_images = str(convert_to_images_raw).lower() in ("true", "1", "yes")

    if not description:
        raise ValueError("Missing required field: description")
    if not rationale:
        raise ValueError("Missing required field: rationale")

    criteria = _build_criterion_prompt(description, rationale)

    log_grader_start(task_id, criteria, is_negative=False)

    try:
        if not input.helper_results:
            raise ValueError("Missing helper results")

        final_answer = input.helper_results[HelperIds.FINAL_ANSWER]
        diff_result = input.helper_results[HelperIds.SNAPSHOT_DIFF]

        model = input.grading_settings.llm_judge_model
        extra_args = input.grading_settings.llm_judge_extra_args

        task_prompt = _extract_task_prompt(input)

        all_artifacts = extract_artifact_changes_from_diff(diff_result)

        log_diff_extraction(task_id, diff_result, all_artifacts, criteria=criteria)

        if not expected_file_type:
            logger.warning(
                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | expected_file_type missing, "
                "defaulting to 'All output'"
            )
            expected_file_type = (
                "All output (modified files and final message in console)"
            )
        elif not is_valid_file_type(expected_file_type):
            logger.warning(
                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | Invalid expected_file_type: "
                f"'{expected_file_type}', defaulting to 'All output'"
            )
            expected_file_type = (
                "All output (modified files and final message in console)"
            )

        filtered_artifacts = _filter_artifacts_programmatically(
            all_artifacts,
            expected_file_type,
            task_id=task_id,
            criteria=criteria,
        )

        if _should_auto_fail_missing_file_type(expected_file_type, filtered_artifacts):
            logger.info(
                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | AUTO-FAIL | "
                f"expected_file_type={expected_file_type} but no matching artifacts"
            )
            return VerifierResult(
                verifier_id=input.verifier.verifier_id,
                verifier_version=input.verifier.verifier_version,
                score=0.0,
                verifier_result_values={
                    "result": 0,
                    "reason": (
                        f"No files matching the expected type ({expected_file_type}) "
                        "were found. The agent did not produce any artifacts of the "
                        "required type."
                    ),
                    "evaluated_artifacts": "",
                    "type": verifier_type,
                    "tags": tags,
                },
            )

        total_artifact_tokens = sum(
            _estimate_artifact_tokens(a, model) for a in filtered_artifacts
        )
        context_limit = get_model_context_limit(model)
        artifact_budget_threshold = int(context_limit * 0.50)

        if total_artifact_tokens <= artifact_budget_threshold:
            logger.info(
                f"[JUDGE][CRITERION_OUTPUT][ARTIFACT_SELECTOR][SKIP] task={task_id} | "
                f"artifacts fit within budget"
            )
            selected_artifacts = filtered_artifacts
        else:
            logger.info(
                f"[JUDGE][CRITERION_OUTPUT][ARTIFACT_SELECTOR][PROCEED] task={task_id} | "
                f"running LLM selection"
            )
            selected_artifacts, _ = await select_artifacts_to_evaluate(
                filtered_artifacts,
                criteria,
                model=model,
                extra_args=extra_args,
                task_id=task_id,
                task_prompt=task_prompt,
            )

        selected_identities = {get_artifact_identity(a) for a in selected_artifacts}
        rejected_artifacts = [
            a
            for a in filtered_artifacts
            if get_artifact_identity(a) not in selected_identities
        ]

        log_artifact_selector_result(
            task_id,
            input_count=len(filtered_artifacts),
            selected_count=len(selected_artifacts),
            selected_artifacts=selected_artifacts,
            criteria=criteria,
            rejected_artifacts=rejected_artifacts if rejected_artifacts else None,
        )

        artifacts_to_reference = None

        if sources:
            parsed_specs = [
                ArtifactSelectionWithTransformations(**spec)
                if isinstance(spec, dict)
                else spec
                for spec in sources
            ]

            input.initial_snapshot_bytes.seek(0)
            with zipfile.ZipFile(input.initial_snapshot_bytes, "r") as initial_zip:
                artifacts_to_reference = await fetch_artifacts_with_transformations(
                    artifacts_to_reference=parsed_specs,
                    initial_snapshot_zip=initial_zip,
                    task_id=task_id,
                    criteria=criteria,
                )
            input.initial_snapshot_bytes.seek(0)

            logger.info(
                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | fetched {len(artifacts_to_reference)} "
                f"reference artifacts from {len(sources)} sources"
            )

        # Convert output artifacts to images for VLM formatting checks
        # Deduplicate by file path to avoid converting multi-sheet xlsx files
        # multiple times (each sheet is a separate artifact with the same path)
        extra_images: list[Any] = []
        if convert_to_images:
            logger.info(
                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | convert_to_images=True, "
                f"converting output artifacts to images"
            )
            converted_paths: set[str] = set()
            input.final_snapshot_bytes.seek(0)
            with zipfile.ZipFile(input.final_snapshot_bytes, "r") as zf:
                for artifact in selected_artifacts:
                    artifact_path = artifact.path
                    if artifact_path in converted_paths:
                        continue
                    ext = Path(artifact_path).suffix.lower()
                    converter = _IMAGE_CONVERTERS.get(ext)
                    if not converter:
                        continue
                    converted_paths.add(artifact_path)

                    # Find the file in the snapshot zip
                    matching = [
                        n
                        for n in zf.namelist()
                        if n.endswith(artifact_path) or artifact_path in n
                    ]
                    if not matching:
                        logger.warning(
                            f"[JUDGE][CRITERION_OUTPUT] task={task_id} | "
                            f"File not found in snapshot: {artifact_path}"
                        )
                        continue

                    file_bytes = zf.read(matching[0])
                    try:
                        result = await converter(file_bytes, Path(artifact_path).name)
                        if result.images:
                            extra_images.extend(result.images)
                            logger.info(
                                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | "
                                f"Converted {artifact_path} to "
                                f"{len(result.images)} image(s)"
                            )
                    except Exception as e:
                        logger.warning(
                            f"[JUDGE][CRITERION_OUTPUT] task={task_id} | "
                            f"Failed to convert {artifact_path}: {e}"
                        )
            input.final_snapshot_bytes.seek(0)

        # Annotate xlsx artifacts with formula info so the LLM can see them
        if selected_artifacts:
            input.final_snapshot_bytes.seek(0)
            with zipfile.ZipFile(input.final_snapshot_bytes, "r") as zf:
                await _annotate_artifacts_with_formulas(selected_artifacts, zf, task_id)
            input.final_snapshot_bytes.seek(0)

        constructed_prompt = build_grading_prompt(
            criteria=criteria,
            final_answer=final_answer,
            model=model,
            artifacts_to_evaluate=selected_artifacts if selected_artifacts else None,
            artifacts_to_reference=artifacts_to_reference,
            include_full_content=True,
            task_id=task_id,
            expected_file_type=expected_file_type,
            task_prompt=task_prompt,
        )

        if constructed_prompt.token_metadata:
            log_grader_truncation(
                task_id,
                was_truncated=constructed_prompt.token_metadata.get(
                    "was_truncated", False
                ),
                original_tokens=constructed_prompt.token_metadata.get(
                    "total_original_tokens", 0
                ),
                final_tokens=constructed_prompt.token_metadata.get(
                    "total_final_tokens", 0
                ),
                files_metadata=constructed_prompt.token_metadata.get("files"),
                criteria=criteria,
            )

        system_prompt = (
            GRADING_SYSTEM_PROMPT
            if artifacts_to_reference
            else GRADING_SYSTEM_PROMPT_NO_REFERENCE
        )

        # Merge converted images with any existing visual artifacts
        all_images = list(constructed_prompt.visual_artifacts_to_evaluate or [])
        if extra_images:
            for i, img in enumerate(extra_images):
                all_images.append(
                    {
                        "type": "converted_image",
                        "url": img.url,
                        "placeholder": f"[CONVERTED_IMAGE_{i + 1}]",
                        "caption": img.caption or f"Converted document page {i + 1}",
                    }
                )
            logger.info(
                f"[JUDGE][CRITERION_OUTPUT] task={task_id} | "
                f"Added {len(extra_images)} converted image(s) to prompt"
            )

        log_grader_final_prompt(
            task_id=task_id,
            criteria=criteria,
            is_negative=False,
            model=model,
            system_prompt_chars=len(system_prompt),
            user_prompt_chars=len(constructed_prompt.user_prompt),
            artifacts_to_evaluate=selected_artifacts if selected_artifacts else None,
            artifacts_to_reference=artifacts_to_reference,
            image_count=len(all_images),
        )

        messages = build_messages(
            system_prompt=system_prompt,
            user_prompt=constructed_prompt.user_prompt,
            images=all_images if all_images else None,
        )
        response = await call_llm(
            model=model,
            messages=messages,
            timeout=LLM_JUDGE_TIMEOUT,
            extra_args=extra_args,
            response_format=GradingResponseSchema,
        )

        choices = response.choices
        if not choices or not isinstance(choices[0], Choices):
            raise ValueError("LLM returned empty response")

        raw_content = choices[0].message.content
        if not raw_content:
            raise ValueError("LLM returned empty content")
        parsed = GradingResponseSchema.model_validate_json(raw_content)

        logger.debug(
            f"[JUDGE][CRITERION_OUTPUT][RESPONSE] task={task_id} | raw_response:\n{raw_content}"
        )

        is_criteria_true = parsed.is_criteria_true
        reason = parsed.rationale

        score = 1.0 if is_criteria_true else 0.0

        evaluated_artifact_names = _get_artifact_display_names(selected_artifacts)

        log_grader_result(
            task_id,
            is_negative=False,
            passed=is_criteria_true,
            score=score,
            criteria=criteria,
        )

        return VerifierResult(
            verifier_id=input.verifier.verifier_id,
            verifier_version=input.verifier.verifier_version,
            score=score,
            verifier_result_values={
                "result": 1 if is_criteria_true else 0,
                "reason": reason,
                "evaluated_artifacts": evaluated_artifact_names,
                "type": verifier_type,
                "tags": tags,
            },
        )

    except Exception as e:
        error_msg = f"criterion_output_verifier grading failed: {str(e)}"
        raise ValueError(error_msg) from e
