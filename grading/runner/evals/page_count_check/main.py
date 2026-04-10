"""Page Count Check eval - verifies file page/slide/sheet counts."""

import fnmatch
import io
import os
import re
import zipfile
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import load_workbook
from pptx import Presentation
from pypdf import PdfReader

from runner.evals.models import EvalImplInput
from runner.models import VerifierResult, VerifierResultStatus

SUPPORTED_EXTENSIONS = {
    ".pdf",  # pages
    ".pptx",  # slides
    ".xlsx",  # sheets
    ".xlsm",  # sheets
}

FILENAME_PATTERN = re.compile(
    r"[\w\-./\\]+\.(?:pdf|pptx|xlsx|xlsm)",
    re.IGNORECASE,
)


def _count_pdf_pages(file_bytes: io.BytesIO) -> int:
    """Count pages in a PDF file."""
    reader = PdfReader(file_bytes)
    return len(reader.pages)


def _count_pptx_slides(file_bytes: io.BytesIO) -> int:
    """Count slides in a PowerPoint file."""
    prs = Presentation(file_bytes)
    return len(prs.slides)


def _count_xlsx_sheets(file_bytes: io.BytesIO) -> int:
    """Count visible sheets in an Excel file."""
    wb = load_workbook(file_bytes, read_only=False)
    visible_count = sum(
        1 for name in wb.sheetnames if wb[name].sheet_state == "visible"
    )
    wb.close()
    return visible_count


def _get_count(file_bytes: io.BytesIO, file_extension: str) -> tuple[int, str]:
    """
    Get page/slide/sheet count for a file.

    Returns:
        Tuple of (count, unit_name) e.g. (9, "pages") or (5, "slides")
    """
    ext = file_extension.lower()

    if ext == ".pdf":
        return _count_pdf_pages(file_bytes), "pages"
    elif ext == ".pptx":
        return _count_pptx_slides(file_bytes), "slides"
    elif ext in (".xlsx", ".xlsm"):
        return _count_xlsx_sheets(file_bytes), "sheets"
    else:
        raise ValueError(f"Unsupported file type for counting: {ext}")


def _find_matching_files(zip_file: zipfile.ZipFile, file_pattern: str) -> list[str]:
    """
    Find files in zip matching the pattern.

    Pattern can be:
    - Exact filename: "report.pdf"
    - Extension pattern: "*.pdf"
    - Path pattern: "output/*.pdf"
    """
    matching = []
    for name in zip_file.namelist():
        if name.endswith("/"):
            continue

        normalized = name.replace("\\", "/").strip("/")
        basename = os.path.basename(normalized)

        if fnmatch.fnmatch(basename, file_pattern) or fnmatch.fnmatch(
            normalized, file_pattern
        ):
            matching.append(name)

    return matching


def _get_supported_extensions() -> set[str]:
    """Get all supported file extensions for counting."""
    return SUPPORTED_EXTENSIONS


def _extract_filename_from_trajectory(input: EvalImplInput) -> str | None:
    """
    Extract a filename from the trajectory messages.

    Searches through all messages looking for file references like:
    - "report.pdf"
    - "output/document.xlsx"
    - "presentation.pptx"

    Returns:
        The first matching filename found, or None if no match.
    """
    if not input.trajectory or not input.trajectory.messages:
        return None

    for msg in input.trajectory.messages:
        content = msg.get("content")
        if not content:
            continue

        # Handle string content
        if isinstance(content, str):
            matches = FILENAME_PATTERN.findall(content)
            if matches:
                # Return the basename of the first match
                filename = os.path.basename(matches[0])
                logger.debug(
                    f"[PAGE_COUNT_CHECK] Extracted filename from trajectory: {filename}"
                )
                return filename

        # Handle list content (multimodal messages)
        elif isinstance(content, list):
            for item in content:
                if isinstance(item, dict) and item.get("type") == "text":
                    text = item.get("text", "")
                    matches = FILENAME_PATTERN.findall(text)
                    if matches:
                        filename = os.path.basename(matches[0])
                        logger.debug(
                            f"[PAGE_COUNT_CHECK] Extracted filename from trajectory: {filename}"
                        )
                        return filename

    return None


async def page_count_check_eval(input: EvalImplInput) -> VerifierResult:
    """
    Verify that files have expected number of pages/slides/sheets.

    Configuration fields:
    - file_pattern: Pattern to match files (e.g., "*.pdf", "output/*.pptx")
    - file_name: Specific filename to check. Overrides file_pattern if provided.
                 Use "{{trajectory}}" to extract filename from conversation messages.
    - min_count: Minimum expected count (inclusive)
    - max_count: Maximum expected count (inclusive)
    - count_mode: "each" (check each file) or "total" (sum all matching files)

    Supported file types:
    - PDF (.pdf): counts pages
    - PowerPoint (.pptx): counts slides
    - Excel (.xlsx, .xlsm): counts visible sheets
    """
    verifier_values = input.verifier.verifier_values or {}
    task_id = input.verifier.task_id or "unknown"

    file_pattern = verifier_values.get("file_pattern", "*.pdf")
    file_name = verifier_values.get("file_name")
    min_count = verifier_values.get("min_count")
    max_count = verifier_values.get("max_count")
    count_mode = verifier_values.get("count_mode", "each")

    # Determine the effective file pattern to use
    if file_name:
        if file_name == "{{trajectory}}":
            # Extract filename from conversation messages
            extracted_filename = _extract_filename_from_trajectory(input)
            if extracted_filename:
                file_pattern = extracted_filename
                logger.info(
                    f"[PAGE_COUNT_CHECK] task={task_id} | Using filename from trajectory: {file_pattern}"
                )
            else:
                logger.warning(
                    f"[PAGE_COUNT_CHECK] task={task_id} | file_name={{{{trajectory}}}} but no filename found, using default pattern: {file_pattern}"
                )
        else:
            # Use the provided filename directly
            file_pattern = file_name
            logger.info(
                f"[PAGE_COUNT_CHECK] task={task_id} | Using specified filename: {file_pattern}"
            )

    if min_count is None and max_count is None:
        raise ValueError("At least one of min_count or max_count must be specified")

    # Preserve original values for reporting, create _val versions for comparisons
    min_count_val = int(min_count) if min_count is not None else 0
    max_count_val = int(max_count) if max_count is not None else float("inf")

    if min_count_val > max_count_val:
        raise ValueError(
            f"Invalid configuration: min_count ({min_count_val}) cannot be greater than max_count ({max_count_val})"
        )

    logger.info(
        f"[PAGE_COUNT_CHECK] task={task_id} | pattern={file_pattern} | "
        f"expected_range=[{min_count_val}, {max_count_val}] | mode={count_mode}"
    )

    try:
        input.final_snapshot_bytes.seek(0)

        with zipfile.ZipFile(input.final_snapshot_bytes, "r") as zf:
            matching_files = _find_matching_files(zf, file_pattern)

            if not matching_files:
                logger.warning(
                    f"[PAGE_COUNT_CHECK] task={task_id} | No files matching '{file_pattern}' found"
                )
                return VerifierResult(
                    verifier_id=input.verifier.verifier_id,
                    verifier_version=input.verifier.verifier_version,
                    score=0.0,
                    verifier_result_values={
                        "passed": False,
                        "reason": f"No files matching pattern '{file_pattern}' found in output",
                        "files_checked": [],
                        "file_pattern": file_pattern,
                        "min_count": min_count,
                        "max_count": max_count if max_count is not None else None,
                    },
                )

            file_counts: list[dict[str, Any]] = []
            # Track counts per unit type to handle mixed file types
            counts_by_unit: dict[str, int] = {}

            for file_path in matching_files:
                try:
                    file_bytes = io.BytesIO(zf.read(file_path))
                    ext = Path(file_path).suffix

                    if ext.lower() not in _get_supported_extensions():
                        logger.warning(
                            f"[PAGE_COUNT_CHECK] task={task_id} | Skipping unsupported file type: {file_path}"
                        )
                        file_counts.append(
                            {
                                "file": file_path,
                                "skipped": True,
                                "reason": f"Unsupported file type: {ext}",
                            }
                        )
                        continue

                    count, unit = _get_count(file_bytes, ext)
                    counts_by_unit[unit] = counts_by_unit.get(unit, 0) + count

                    file_counts.append(
                        {
                            "file": file_path,
                            "count": count,
                            "unit": unit,
                        }
                    )

                    logger.info(
                        f"[PAGE_COUNT_CHECK] task={task_id} | file={file_path} | {unit}={count}"
                    )

                except Exception as e:
                    logger.error(
                        f"[PAGE_COUNT_CHECK] task={task_id} | Failed to count in {file_path}: {e}"
                    )
                    file_counts.append(
                        {
                            "file": file_path,
                            "error": str(e),
                        }
                    )

        input.final_snapshot_bytes.seek(0)

        successfully_counted = sum(
            1 for fc in file_counts if "count" in fc and "skipped" not in fc
        )

        # If no files were successfully counted, fail the verification
        if successfully_counted == 0:
            skipped_count = sum(1 for fc in file_counts if fc.get("skipped"))
            error_count = sum(1 for fc in file_counts if "error" in fc)

            if error_count > 0:
                reason = f"All {len(file_counts)} file(s) failed to process ({error_count} errors)"
            elif skipped_count > 0:
                reason = f"All {len(file_counts)} file(s) have unsupported extensions and were skipped"
            else:
                reason = "No files were successfully counted"

            logger.warning(f"[PAGE_COUNT_CHECK] task={task_id} | {reason}")
            return VerifierResult(
                verifier_id=input.verifier.verifier_id,
                verifier_version=input.verifier.verifier_version,
                score=0.0,
                verifier_result_values={
                    "passed": False,
                    "reason": reason,
                    "files_checked": file_counts,
                    "file_pattern": file_pattern,
                    "min_count": min_count,
                    "max_count": max_count if max_count is not None else None,
                },
            )

        if max_count_val == float("inf"):
            range_str = f">= {min_count_val}"
        elif min_count_val == 0:
            range_str = f"<= {max_count_val}"
        else:
            range_str = f"{min_count_val}-{max_count_val}"

        unique_units = list(counts_by_unit.keys())
        if len(unique_units) == 0:
            unit_name = "items"
        elif len(unique_units) == 1:
            unit_name = unique_units[0]
        else:
            unit_name = "items"

        total_count = sum(counts_by_unit.values())

        if count_mode == "total":
            if len(unique_units) > 1:
                breakdown = " + ".join(
                    f"{count} {unit}" for unit, count in sorted(counts_by_unit.items())
                )
                passed = min_count_val <= total_count <= max_count_val
                reason = (
                    f"Total items: {total_count} ({breakdown}) (expected: {range_str}) - "
                    f"{'PASS' if passed else 'FAIL'}"
                )
            else:
                passed = min_count_val <= total_count <= max_count_val
                reason = (
                    f"Total {unit_name}: {total_count} (expected: {range_str}) - "
                    f"{'PASS' if passed else 'FAIL'}"
                )

            return VerifierResult(
                verifier_id=input.verifier.verifier_id,
                verifier_version=input.verifier.verifier_version,
                score=1.0 if passed else 0.0,
                verifier_result_values={
                    "passed": passed,
                    "reason": reason,
                    "total_count": total_count,
                    "counts_by_unit": counts_by_unit,
                    "unit": unit_name,
                    "files_checked": file_counts,
                    "file_pattern": file_pattern,
                    "min_count": min_count,
                    "max_count": max_count if max_count is not None else None,
                    "count_mode": count_mode,
                },
            )

        else:
            all_passed = True
            failed_files = []

            for fc in file_counts:
                if "error" in fc:
                    all_passed = False
                    failed_files.append(fc["file"])
                elif "skipped" in fc:
                    continue
                elif not (min_count_val <= fc["count"] <= max_count_val):
                    all_passed = False
                    failed_files.append(f"{fc['file']} ({fc['count']} {fc['unit']})")

            if all_passed:
                checked_count = sum(
                    1 for fc in file_counts if "count" in fc and "skipped" not in fc
                )
                if len(unique_units) > 1:
                    # Mixed file types
                    reason = f"All {checked_count} file(s) have counts within expected range ({range_str})"
                else:
                    reason = f"All {checked_count} file(s) have {unit_name} within expected range ({range_str})"
            else:
                reason = f"Files outside expected range ({range_str}): {', '.join(failed_files)}"

            return VerifierResult(
                verifier_id=input.verifier.verifier_id,
                verifier_version=input.verifier.verifier_version,
                score=1.0 if all_passed else 0.0,
                verifier_result_values={
                    "passed": all_passed,
                    "reason": reason,
                    "counts_by_unit": counts_by_unit,
                    "files_checked": file_counts,
                    "file_pattern": file_pattern,
                    "min_count": min_count,
                    "max_count": max_count if max_count is not None else None,
                    "count_mode": count_mode,
                },
            )

    except Exception as e:
        logger.error(f"[PAGE_COUNT_CHECK] task={task_id} | Error: {e}")
        return VerifierResult(
            verifier_id=input.verifier.verifier_id,
            verifier_version=input.verifier.verifier_version,
            score=0.0,
            verifier_result_values={
                "passed": False,
                "error": str(e),
            },
            status=VerifierResultStatus.ERROR,
            message=str(e),
        )
